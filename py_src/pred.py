# coding=utf-8
import os
import sys
import numpy as np
from sklearn.neighbors import KNeighborsClassifier
from sklearn.naive_bayes import GaussianNB
from sklearn.linear_model import LogisticRegression
from sklearn import svm,metrics
from sklearn.tree import DecisionTreeClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.externals import joblib
import readData
import openpyxl
def cos_sim(vector_a, vector_b):
    vector_a = np.mat(vector_a)
    vector_b = np.mat(vector_b)
    num = float(vector_a * vector_b.T)
    denom = np.linalg.norm(vector_a) * np.linalg.norm(vector_b)
    cos = num / denom
    # normalization to [0,1]
    sim = 0.5 + 0.5 * cos
    return sim

def load_pred(model,x_test,len_data_test,sheet1,line_num,method):
    clf = joblib.load(model)
    clf_pred = clf.predict(x_test)
    clf_proba = clf.predict_proba(x_test)

    sheet1.cell(line_num+1, 0+1).value= method
    for jj in range(len_data_test):
        sheet1.cell(line_num+1, jj+1+1).value=(str(clf_pred[jj]) + '(' + str(round(max(clf_proba[jj]), 3)) + ')')
    line_num += 1
    return sheet1,line_num

def load_compute(model,data_input):
    if (not os.path.exists(model)):
        print(model+'does not exist')
        return
    clf = joblib.load(model)
    clf_proba = clf.predict_proba(data_input)
    return clf_proba

def train(train_path):
    len_train_file = len(os.listdir(train_path))
    dirs = os.listdir(train_path)

    for file_num in range(len_train_file):
        train_file = os.path.join(train_path,dirs[file_num])
        save_model_path ='output/train_model/'+ train_path.split('\\')[-1]+'/'
        if not os.path.exists(save_model_path):
            os.makedirs(save_model_path)
        alleleGroupName, alleleList, alleleData,data,groupName,shuffleIndex = readData.readExcelData(train_file, SHUFFLE=True)

        x_train = alleleData
        y_train = alleleGroupName

        #knn
        knn = KNeighborsClassifier(1).fit(x_train, y_train)
        joblib.dump(knn, save_model_path+dirs[file_num].split('.')[0]+'_knn.model')

        #naiveBayes
        gnb = GaussianNB().fit(x_train, y_train)
        joblib.dump(gnb, save_model_path+dirs[file_num].split('.')[0]+'_naiveBayes.model')

        #logisticRegression
        lr = LogisticRegression(penalty="l2", C=1, multi_class="ovr", solver="newton-cg",max_iter=1000).fit(x_train, y_train)
        joblib.dump(lr, save_model_path+dirs[file_num].split('.')[0]+'_logisticRegression.model')

        #SVM
        svm_svm = svm.SVC(probability = True).fit(x_train, y_train)
        joblib.dump(svm_svm, save_model_path+dirs[file_num].split('.')[0]+'_svm.model')

        #decesionTree
        dt = DecisionTreeClassifier(criterion='entropy').fit(x_train, y_train)
        joblib.dump(dt, save_model_path+dirs[file_num].split('.')[0]+'_decesionTree.model')

        #randomForest
        rf = RandomForestClassifier(n_estimators=170).fit(x_train, y_train)
        joblib.dump(rf, save_model_path+dirs[file_num].split('.')[0]+'_randomForest.model')

def test(mode,train_path,test_file):
    ml_method = ['knn','naiveBayes','logisticRegression','svm','decesionTree','randomForest']
    len_train_file = len(os.listdir(train_path)) if mode=="others" else 1
    dirs = os.listdir(train_path)
    if(mode=="DATABASE"):
        save_model_path=train_path+'/'
        save_result = 'output/PredictResult_DATABASE_'+test_file.split('\\')[-1].split('.')[0]+'.xlsx'

    else:
        save_model_path ='output/train_model/'+ train_path.split('\\')[-1]+'/'
        if not os.path.exists(save_model_path):
            os.makedirs(save_model_path)
        save_result = 'output/PredictResult_'+ train_path.split('\\')[-1]+'_'+test_file.split('\\')[-1].split('.')[0]+'.xlsx'
    try:
        file = openpyxl.Workbook()
        sheet1 = file.create_sheet('predict')
    except IOError:
        return "Error: Failed to create file"

    sampleName_test,alleleGroupName_test, alleleList_test, data_test= readData.readExcelData(test_file, SHUFFLE=False)

    line_num = 0
    for file_num in range(len_train_file):
        sheet1.cell(line_num+1, 0+1).value=dirs[file_num] if mode=="others" else "DATABASE"
        line_num+=1
        for data_num in range(len(data_test)):
            sheet1.cell(line_num+1, data_num+1+1).value=str(sampleName_test[data_num])
        line_num+=1
        x_test = data_test

        for k in range(6):
            if(mode=="others"):
                model = os.path.join(save_model_path,dirs[file_num].split('.')[0]+'_'+ml_method[k] + '.model')
            else:
                model=os.path.join(save_model_path,ml_method[k]+'.model')

            try:
                fout = open(model)
            except:
                print(model + ' do not exist. Please train first.')
                return
            sheet1, line_num = load_pred( model, x_test, len(data_test), sheet1, line_num,
                                         ml_method[k])

        # #knn
        # model = os.path.join(save_model_path,dirs[file_num].split('.')[0]+'_knn.model')
        #
        # #naiveBayes
        # model = os.path.join(save_model_path,dirs[file_num].split('.')[0]+'_naiveBayes.model')
        # sheet1, line_num=load_pred(mode,model,train_path,x_test,len(data_test),sheet1,line_num,"naiveBayes")
        #
        # #logisticRegression
        # model = os.path.join(save_model_path, dirs[file_num].split('.')[0] + '_logisticRegression.model')
        # sheet1, line_num=load_pred(mode,model,train_path,x_test,len(data_test),sheet1,line_num,"logisticRegression")
        #
        # #SVM
        # model = os.path.join(save_model_path, dirs[file_num].split('.')[0] + '_svm.model')
        # sheet1, line_num=load_pred(mode,model,train_path,x_test,len(data_test),sheet1,line_num,"svm")
        #
        # #decesionTree
        # model = os.path.join(save_model_path, dirs[file_num].split('.')[0] + '_decesionTree.model')
        # sheet1, line_num=load_pred(mode,model,train_path,x_test,len(data_test),sheet1,line_num,"decesionTree")
        #
        # #randomForest
        # model = os.path.join(save_model_path, dirs[file_num].split('.')[0] + '_randomForest.model')
        # sheet1, line_num=load_pred(mode,model,train_path,x_test,len(data_test),sheet1,line_num,"randomForest")
    try:
        file.save(save_result)
    except IOError:
        return "Error: File not found or failed to read "
    else:
        print(save_result+" has been written.")
        return save_result+" has been written."

def compute(train_path,compared_file,input_file):
    ml_method = ['knn','naiveBayes','logisticRegression','svm','decesionTree','randomForest']
    save_result = 'output//SimilarityResult_DATABASE_'+ compared_file.split('\\')[-1].split('.')[0]+'_'+input_file.split('\\')[-1].split('.')[0]+'.xlsx'
    try:
        file = openpyxl.Workbook()
        sheet1 = file.create_sheet('similarity')
    except IOError:
        return "Error: Failed to create file"

    sampleName_input, _, alleleList_input, data_input= readData.readExcelData(input_file, SHUFFLE=False)
    sampleName_compared, _, alleleList_compared, data_compared= readData.readExcelData(compared_file, SHUFFLE=False)

    proba_compared= list()
    proba_input = list()

    for k in range(6):
        model = os.path.join(train_path, ml_method[k]+'.model')
        proba_input.append(load_compute(model, data_input))
        proba_compared.append(load_compute(model, data_compared))

    line_num=1
    sheet1.cell(line_num, 1).value = 'method'
    for k in range(6):
        sheet1.cell(line_num+1, 1).value = ml_method[k]
        for i in range(len(data_compared)):
            sheet1.cell(line_num+1, 2).value= sampleName_compared[i]
            line_num+=1
            for j in range(len(data_input)):
                sheet1.cell(1, j+3).value= sampleName_input[j]
                sheet1.cell(line_num, j+3).value= str(round(cos_sim(proba_compared[k][i],proba_input[k][j]), 5))
    try:
        file.save(save_result)
    except IOError:
        return "Error: File not found or failed to read "
    else:
        print(save_result+" has been written.")
        return save_result+" has been written."

if __name__ == "__main__":
    # train(".\\20181021")
    # test("DATABASE","E:\\Projects\\mengyuan_code_release\\predict_new_data\\trained_DATABASE_model","input1.xls")
    # test("others",".\\20181021","input1.xls")
    # compute("E:\\projects_java\\trained_DATABASE_model",
    #         "E:\\projects_java\\src\\input1.xls","E:\\projects_java\\src\\input1.xls")

    if(sys.argv[1]=="train"):
        train(sys.argv[2])
    elif(sys.argv[1]=="test"):
        test(sys.argv[2],sys.argv[3],sys.argv[4])
    elif(sys.argv[1]=="compute"):
        compute(sys.argv[2],sys.argv[3],sys.argv[4])